from __future__ import annotations

import argparse
import base64
import copy
import io
import json
import re
import sys
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

PROJECT_ROOT = Path(__file__).resolve().parents[3]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from scripts.e2e.adapters.security_tool.action_templates import ACTION_TEMPLATES
from scripts.e2e.adapters.security_tool.config import ADAPTER_CONFIG
from scripts.e2e.adapters.security_tool.flows.registry import FLOW_REGISTRY
from scripts.e2e.adapters.security_tool.resolvers import PAGE_REGISTRY
from scripts.e2e.core.resolver_contracts import LOCAL_ONLY_FLOW_REFS
from scripts.e2e.adapters.security_tool.suites import SUITES
from scripts.e2e.core.contracts import validate_case_contract

METADATA_DIR = PROJECT_ROOT / "scripts/e2e/metadata"
CASES_DIR = PROJECT_ROOT / "scripts/e2e/cases"
RESULTS_DIR = PROJECT_ROOT / "scripts/e2e/results"
CATALOG_PATH = METADATA_DIR / "case_catalog.json"
COVERAGE_PATH = METADATA_DIR / "coverage_snapshot.json"
IMPORT_REPORT_PATH = METADATA_DIR / "import_report.json"
VALIDATION_REPORT_PATH = METADATA_DIR / "validation_report.json"
BRIDGE_MAP_PATH = PROJECT_ROOT / "scripts/e2e/tools/bridge_action_map.json"

CATALOG_VERSION = "1.0"
XML_NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

EDITABLE_FIELDS = {
    "preconditions_text",
    "steps_text",
    "checkpoints_text",
    "notes",
    "suite_membership",
    "review_status",
    "review_notes",
    "status",
    "source",
    "capability_gap_status",
    "capability_gap_items",
    "bridge_coverage_status",
}

MODULE_NAME_TO_ID = {
    "安全总览": "dashboard",
    "dashboard": "dashboard",
    "防火墙": "firewall",
    "防火墙管理": "firewall",
    "firewall": "firewall",
    "日志管理": "log-manage",
    "log-manage": "log-manage",
    "logs": "log-manage",
    "外设管理": "peripheral",
    "peripheral": "peripheral",
    "peripheral-manage": "peripheral",
    "身份鉴别": "identity",
    "identity": "identity",
    "工具设置": "tool-settings",
    "tool-settings": "tool-settings",
    "tool_settings": "tool-settings",
    "navigation": "navigation",
    "bootstrap": "bootstrap",
}

MODULE_ID_TO_NAME = {
    "dashboard": "安全总览",
    "firewall": "防火墙管理",
    "log-manage": "日志管理",
    "peripheral": "外设管理",
    "identity": "身份鉴别",
    "tool-settings": "工具设置",
    "navigation": "导航",
    "bootstrap": "启动初始化",
}

MODULE_ID_TO_PAGE_ID = {
    "dashboard": "dashboard",
    "firewall": "firewall",
    "log-manage": "log-manage",
    "peripheral": "peripheral-manage",
    "identity": "identity",
    "tool-settings": "tool-settings",
}

HEADER_ALIASES = {
    "case_id": {"用例编号", "case_id", "case id", "编号"},
    "case_name": {"用例名称", "case_name", "名称", "功能"},
    "module_name": {"所属模块", "模块", "module", "场景"},
    "enabled": {"是否启用", "enabled", "启用"},
    "preconditions_text": {"前置条件", "preconditions", "前置", "场景描述"},
    "steps_text": {"操作步骤", "步骤", "测试步骤", "功能详细描述", "steps"},
    "checkpoints_text": {"检查点", "预期结果", "测试用例", "expected", "checkpoints"},
    "notes": {"备注", "notes", "大致接口"},
}


@dataclass
class ImportResult:
    catalog: dict[str, Any]
    report: dict[str, Any]


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def ensure_directories() -> None:
    METADATA_DIR.mkdir(parents=True, exist_ok=True)


def to_posix(path: Path) -> str:
    return str(path.relative_to(PROJECT_ROOT)).replace("\\", "/")


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return copy.deepcopy(default)
    return json.loads(path.read_text(encoding="utf-8"))


def save_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def default_catalog() -> dict[str, Any]:
    return {"version": CATALOG_VERSION, "updated_at": utc_now(), "records": []}


def load_catalog() -> dict[str, Any]:
    catalog = load_json(CATALOG_PATH, default_catalog())
    if "records" not in catalog or not isinstance(catalog["records"], list):
        catalog = default_catalog()
    catalog.setdefault("version", CATALOG_VERSION)
    catalog.setdefault("updated_at", utc_now())
    return catalog


def save_catalog(catalog: dict[str, Any]) -> dict[str, Any]:
    ensure_directories()
    records = sorted(catalog.get("records", []), key=lambda item: item.get("case_id", ""))
    payload = {"version": catalog.get("version", CATALOG_VERSION), "updated_at": utc_now(), "records": records}
    save_json(CATALOG_PATH, payload)
    return payload


def load_bridge_action_map() -> dict[str, str]:
    return load_json(BRIDGE_MAP_PATH, {})


def normalize_text(value: Any) -> str:
    text = str(value or "").strip()
    return re.sub(r"\s+", " ", text)


def module_id_to_name(module_id: str) -> str:
    return MODULE_ID_TO_NAME.get(module_id, module_id)


def normalize_module_id(module_name: str) -> str:
    key = normalize_text(module_name).lower()
    if not key:
        return ""
    if key in MODULE_NAME_TO_ID:
        return MODULE_NAME_TO_ID[key]
    for candidate, module_id in MODULE_NAME_TO_ID.items():
        if candidate.lower() == key:
            return module_id
    return ""


def module_to_page_id(module_id: str) -> str:
    return MODULE_ID_TO_PAGE_ID.get(module_id, module_id)


def default_record(case_id: str = "") -> dict[str, Any]:
    return {
        "case_id": case_id,
        "case_name": "",
        "module_id": "",
        "module_name": "",
        "status": "draft",
        "source": "generated",
        "preconditions_text": "",
        "steps_text": "",
        "checkpoints_text": "",
        "notes": [],
        "structured_flow": [],
        "structured_assertions": [],
        "checkpoint_summary": [],
        "suite_membership": [],
        "bridge_coverage_status": "missing",
        "capability_gap_status": "none",
        "capability_gap_items": [],
        "capability_gap_breakdown": {
            "flow": [],
            "template": [],
            "params": [],
            "assertion": [],
            "bridge": [],
            "module": [],
            "other": [],
        },
        "case_path": "",
        "last_result_status": "NOT_RUN",
        "review_status": "unreviewed",
        "review_notes": [],
    }


def slugify(text: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9]+", "_", text.strip().lower()).strip("_")
    return cleaned or "step"


def default_flow_name(flow_ref: str, index: int) -> str:
    return f"{index + 1:02d}_{slugify(flow_ref.replace('.', '_'))}"


def normalize_flow_items(flow_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    for index, item in enumerate(flow_items):
        ref = item.get("ref") or item.get("action", "")
        params = item.get("params", {})
        normalized.append(
            {
                "name": item.get("name") or default_flow_name(ref, index),
                "ref": ref,
                "params": params if isinstance(params, dict) else {},
            }
        )
    return normalized


def unique_list(items: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for item in items:
        value = normalize_text(item)
        if not value or value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered


def load_result_status_map() -> dict[str, dict[str, str]]:
    status_map: dict[str, dict[str, str]] = {}
    for result_path in RESULTS_DIR.glob("*.json"):
        if result_path.name.endswith("_suite_summary.json"):
            continue
        try:
            payload = json.loads(result_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            continue
        case_id = payload.get("case_id")
        finished_at = payload.get("finished_at", "")
        if not case_id:
            continue
        current = status_map.get(case_id)
        if current is None or finished_at >= current.get("finished_at", ""):
            status_map[case_id] = {"status": payload.get("status", "NOT_RUN"), "finished_at": finished_at}
    return status_map


def infer_suite_membership(case_relative_path: str) -> list[str]:
    memberships: list[str] = []
    for suite_name, suite_cases in SUITES.items():
        if case_relative_path in suite_cases:
            memberships.append(suite_name)
    return memberships


def extract_checkpoint_summary(case_payload: dict[str, Any]) -> list[str]:
    summary: list[str] = []
    for note in case_payload.get("notes", []):
        if "evidence" in note.lower() or "verify" in note.lower() or "状态" in note:
            summary.append(note)
    for assertion in case_payload.get("assertions", []):
        value = assertion.get("value")
        if value:
            summary.append(str(value))
    return unique_list(summary)


def derive_template_key(flow_ref: str, params: dict[str, Any]) -> str:
    if not flow_ref.startswith("entity."):
        return ""
    domain = str(params.get("domain", "")).strip()
    entity = str(params.get("entity", "")).strip()
    action = str(params.get("action", flow_ref.split(".")[-1])).strip()
    variant = str(params.get("variant", "")).strip().lower()
    return ".".join(part for part in (domain, entity, action, variant) if part)


def collect_template_placeholders(template: dict[str, Any]) -> set[str]:
    placeholders: set[str] = set()

    def visit(value: Any) -> None:
        if isinstance(value, str):
            for match in re.findall(r"\$\{([^}]+)\}", value):
                placeholders.add(match)
            return
        if isinstance(value, list):
            for item in value:
                visit(item)
            return
        if isinstance(value, dict):
            for item in value.values():
                visit(item)

    visit(template)
    return placeholders


TEMPLATE_REQUIRED_DATA = {
    "firewall.rule.create.domain": ["domain", "direction", "policy"],
    "firewall.rule.delete.domain": ["domain"],
    "tool_settings.startup_auth.toggle": ["target_state"],
    "tool_settings.auth_method.update": ["method"],
    "peripheral.interface.toggle.usb": ["target_state"],
    "peripheral.interface.toggle.bluetooth": ["target_state"],
    "peripheral.interface.toggle.wifi": ["target_state"],
    "peripheral.interface.toggle.hdc": ["target_state"],
    "peripheral.usb_storage_policy.update": ["policy"],
    "peripheral.whitelist.create.usb": ["device_id"],
    "peripheral.whitelist.create.bluetooth": ["device_id"],
    "peripheral.blacklist.create.usb": ["device_id"],
    "identity.password_policy.update": ["min_length"],
    "identity.domain_policy.update": ["password_max_age_days", "expiration_notify_days", "auth_validity_minutes"],
}


def template_support_issues(flow_ref: str, params: dict[str, Any]) -> list[str]:
    if not flow_ref.startswith("entity."):
        return []
    template_key = derive_template_key(flow_ref, params)
    if not template_key:
        return ["缺少 template_key 组成字段"]
    template = ACTION_TEMPLATES.get(template_key)
    if not template:
        return [f"模板未定义: {template_key}"]

    data = params.get("data", {}) if isinstance(params.get("data", {}), dict) else {}
    issues: list[str] = []
    required_keys = TEMPLATE_REQUIRED_DATA.get(template_key)
    if required_keys is None:
        required_keys = sorted(
            placeholder.split(".", 1)[1]
            for placeholder in collect_template_placeholders(template)
            if placeholder.startswith("data.")
        )
    for key in required_keys:
        value = data.get(key)
        if value in ("", None):
            issues.append(f"模板缺少必填数据: {template_key}.{key}")
    return issues


def compute_bridge_status(flow_items: list[Any], bridge_map: dict[str, str]) -> str:
    if not flow_items:
        return "missing"
    statuses = []
    for item in flow_items:
        flow_ref = item.get("ref", "") if isinstance(item, dict) else str(item)
        params = item.get("params", {}) if isinstance(item, dict) else {}
        if flow_ref in LOCAL_ONLY_FLOW_REFS:
            statuses.append("covered")
            continue
        if flow_ref not in FLOW_REGISTRY:
            statuses.append("missing")
            continue
        if not bridge_map.get(flow_ref):
            statuses.append("missing")
            continue
        if flow_ref.startswith("entity."):
            statuses.append("covered" if not template_support_issues(flow_ref, params) else "missing")
        else:
            statuses.append("covered")
    if all(status == "covered" for status in statuses):
        return "covered"
    if any(status == "covered" for status in statuses):
        return "partial"
    return "missing"


def build_gap_breakdown(gap_items: list[str]) -> dict[str, list[str]]:
    breakdown = {
        "flow": [],
        "template": [],
        "params": [],
        "assertion": [],
        "bridge": [],
        "module": [],
        "other": [],
    }
    for item in unique_list(gap_items):
        lower_item = item.lower()
        if "unknown flow" in lower_item or "flow" in lower_item or "未知 flow" in item or "缺少可执行 flow" in item:
            bucket = "flow"
        elif "template" in lower_item or "模板未定义" in item:
            bucket = "template"
        elif "required data" in lower_item or "param" in lower_item or "模板缺少必填数据" in item:
            bucket = "params"
        elif "assertion" in lower_item or "断言" in item:
            bucket = "assertion"
        elif "bridge" in lower_item or "Bridge" in item:
            bucket = "bridge"
        elif "module" in lower_item or "模块" in item:
            bucket = "module"
        else:
            bucket = "other"
        breakdown[bucket].append(item)
    return breakdown


def classify_gap_status(flow_items: list[dict[str, Any]], assertions: list[dict[str, Any]], bridge_status: str) -> tuple[str, list[str]]:
    gaps: list[str] = []
    if not flow_items:
        gaps.append("缺少可执行 flow")
    for item in flow_items:
        ref = item.get("ref", "")
        if ref and ref not in FLOW_REGISTRY:
            gaps.append(f"未知 flow 引用: {ref}")
        gaps.extend(template_support_issues(ref, item.get("params", {})))
    if bridge_status != "covered":
        gaps.append("Bridge 覆盖未完整")
    if not assertions:
        gaps.append("缺少断言")
    if not gaps:
        return "none", []
    if any("未知 flow" in gap or "缺少可执行 flow" in gap for gap in gaps):
        return ("flow_gap" if bridge_status == "covered" else "mixed"), gaps
    if bridge_status != "covered":
        return ("bridge_gap" if assertions else "mixed"), gaps
    return "assertion_gap", gaps


def build_case_catalog() -> dict[str, Any]:
    ensure_directories()
    existing_catalog = load_catalog()
    existing_records = {record.get("case_id"): record for record in existing_catalog.get("records", [])}
    result_status_map = load_result_status_map()
    bridge_map = load_bridge_action_map()
    records: list[dict[str, Any]] = []

    for case_path in sorted(CASES_DIR.rglob("*.json")):
        payload = json.loads(case_path.read_text(encoding="utf-8"))
        case_id = payload.get("case_id")
        if not case_id:
            continue
        relative_case_path = to_posix(case_path)
        suite_membership = infer_suite_membership(str(case_path.relative_to(CASES_DIR)).replace("\\", "/"))
        flow_items = normalize_flow_items(payload.get("flow", []))
        bridge_status = compute_bridge_status(flow_items, bridge_map)
        gap_status, gap_items = classify_gap_status(flow_items, payload.get("assertions", []), bridge_status)

        record = default_record(case_id)
        previous = existing_records.get(case_id, {})
        for key in EDITABLE_FIELDS:
            if key in previous:
                record[key] = copy.deepcopy(previous[key])

        record.update(
            {
                "case_id": case_id,
                "case_name": payload.get("case_name", ""),
                "module_id": payload.get("module", ""),
                "module_name": module_id_to_name(payload.get("module", "")),
                "status": previous.get("status", "implemented") if previous.get("status") in {"implemented", "draft", "planned", "deprecated"} else "implemented",
                "source": previous.get("source", "manual"),
                "structured_flow": flow_items,
                "structured_assertions": payload.get("assertions", []),
                "checkpoint_summary": previous.get("checkpoint_summary") or extract_checkpoint_summary(payload),
                "suite_membership": suite_membership,
                "bridge_coverage_status": bridge_status,
                "capability_gap_status": gap_status,
                "capability_gap_items": unique_list(previous.get("capability_gap_items", []) + gap_items),
                "capability_gap_breakdown": build_gap_breakdown(unique_list(previous.get("capability_gap_items", []) + gap_items)),
                "case_path": relative_case_path,
                "last_result_status": result_status_map.get(case_id, {}).get("status", "NOT_RUN"),
            }
        )
        records.append(record)

    catalog = {"version": CATALOG_VERSION, "updated_at": utc_now(), "records": sorted(records, key=lambda item: item.get("case_id", ""))}
    return save_catalog(catalog)


def build_coverage_snapshot(catalog: dict[str, Any] | None = None) -> dict[str, Any]:
    catalog = catalog or load_catalog()
    records = catalog.get("records", [])
    bridge_map = load_bridge_action_map()
    module_stats: dict[str, dict[str, int]] = {}
    suite_stats: dict[str, int] = {}
    flow_stats: dict[str, int] = {}
    bridge_stats: dict[str, int] = {}
    result_stats: dict[str, int] = {}
    status_stats: dict[str, int] = {}

    for record in records:
        module_id = record.get("module_id", "")
        module_stats.setdefault(module_id, {"total": 0, "implemented": 0, "draft": 0, "planned": 0, "deprecated": 0})
        module_stats[module_id]["total"] += 1
        module_stats[module_id][record.get("status", "draft")] = module_stats[module_id].get(record.get("status", "draft"), 0) + 1
        status_stats[record.get("status", "draft")] = status_stats.get(record.get("status", "draft"), 0) + 1
        result_stats[record.get("last_result_status", "NOT_RUN")] = result_stats.get(record.get("last_result_status", "NOT_RUN"), 0) + 1

        for suite_name in record.get("suite_membership", []):
            suite_stats[suite_name] = suite_stats.get(suite_name, 0) + 1

        for flow_item in record.get("structured_flow", []):
            flow_ref = flow_item.get("ref", "")
            if not flow_ref:
                continue
            flow_stats[flow_ref] = flow_stats.get(flow_ref, 0) + 1
            action = bridge_map.get(flow_ref, "")
            if action:
                bridge_stats[action] = bridge_stats.get(action, 0) + 1

    snapshot = {
        "updated_at": utc_now(),
        "total_cases": len(records),
        "status_stats": status_stats,
        "result_stats": result_stats,
        "module_stats": module_stats,
        "suite_stats": suite_stats,
        "flow_stats": flow_stats,
        "bridge_action_stats": bridge_stats,
        "bridge_coverage": {
            "covered": sum(1 for record in records if record.get("bridge_coverage_status") == "covered"),
            "partial": sum(1 for record in records if record.get("bridge_coverage_status") == "partial"),
            "missing": sum(1 for record in records if record.get("bridge_coverage_status") == "missing"),
        },
    }
    save_json(COVERAGE_PATH, snapshot)
    return snapshot


def xlsx_column_index(cell_ref: str) -> int:
    letters = "".join(char for char in cell_ref if char.isalpha())
    index = 0
    for char in letters:
        index = index * 26 + (ord(char.upper()) - ord("A") + 1)
    return max(index - 1, 0)


def parse_xlsx_rows(data: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        worksheet = workbook.active
        raw_rows = [
            [normalize_text(cell) for cell in row]
            for row in worksheet.iter_rows(values_only=True)
        ]
        raw_rows = [row for row in raw_rows if any(normalize_text(value) for value in row)]
        if raw_rows:
            header = raw_rows[0]
            rows: list[dict[str, str]] = []
            for row_values in raw_rows[1:]:
                width = max(len(header), len(row_values))
                row = {
                    normalize_text(header[index]) if index < len(header) else f"column_{index}":
                    normalize_text(row_values[index]) if index < len(row_values) else ""
                    for index in range(width)
                }
                if any(normalize_text(value) for value in row.values()):
                    rows.append(row)
            if rows:
                return rows
    except Exception:
        pass

    rows: list[dict[str, str]] = []
    with zipfile.ZipFile(io.BytesIO(data)) as workbook:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in workbook.namelist():
            shared_root = ET.fromstring(workbook.read("xl/sharedStrings.xml"))
            for node in shared_root.findall("a:si", XML_NS):
                shared_strings.append("".join(text.text or "" for text in node.iterfind(".//a:t", XML_NS)))

        first_sheet = "xl/worksheets/sheet1.xml"
        if first_sheet not in workbook.namelist():
            sheet_entry = next((name for name in workbook.namelist() if name.startswith("xl/worksheets/sheet")), "")
            if not sheet_entry:
                return []
            first_sheet = sheet_entry

        sheet_root = ET.fromstring(workbook.read(first_sheet))
        raw_rows: list[list[str]] = []
        for row in sheet_root.findall(".//a:sheetData/a:row", XML_NS):
            cells: dict[int, str] = {}
            for cell in row.findall("a:c", XML_NS):
                index = xlsx_column_index(cell.get("r", "A1"))
                cell_type = cell.get("t", "")
                value_node = cell.find("a:v", XML_NS)
                value = "" if value_node is None else value_node.text or ""
                if cell_type == "s" and value:
                    value = shared_strings[int(value)]
                elif cell_type == "inlineStr":
                    value = "".join(text.text or "" for text in cell.iterfind(".//a:t", XML_NS))
                cells[index] = normalize_text(value)
            if not cells:
                continue
            width = max(cells) + 1
            raw_rows.append([cells.get(index, "") for index in range(width)])

    if not raw_rows:
        return []

    header = raw_rows[0]
    normalized_headers = [normalize_text(item) for item in header]
    for row_values in raw_rows[1:]:
        row = {normalized_headers[index]: row_values[index] if index < len(row_values) else "" for index in range(len(normalized_headers))}
        if any(normalize_text(value) for value in row.values()):
            rows.append(row)
    return rows


def resolve_header_value(row: dict[str, str], field_name: str) -> str:
    aliases = HEADER_ALIASES[field_name]
    alias_set = {alias.lower() for alias in aliases}
    for header, value in row.items():
        if normalize_text(header).lower() in alias_set:
            return normalize_text(value)
    return ""


def derive_case_id(module_id: str, row_index: int, case_name: str) -> str:
    prefix_map = {
        "dashboard": "DASH",
        "firewall": "FW",
        "log-manage": "LOG",
        "peripheral": "PER",
        "identity": "ID",
        "tool-settings": "SET",
        "navigation": "NAV",
        "bootstrap": "BOOT",
    }
    prefix = prefix_map.get(module_id, "AUTO")
    suffix = row_index + 1
    if case_name:
        name_slug = slugify(case_name).upper()[:8]
        if name_slug:
            return f"{prefix}-{name_slug}-{suffix:03d}"
    return f"{prefix}-AUTO-{suffix:03d}"


def split_text_sentences(text: str) -> list[str]:
    raw_parts = re.split(r"[；;。\n]+", text)
    return unique_list([part.strip(" 1234567890.、") for part in raw_parts if normalize_text(part)])


def summarize_checkpoints(checkpoints_text: str, steps_text: str) -> list[str]:
    summary = split_text_sentences(checkpoints_text)
    if summary:
        return summary
    inferred = []
    for sentence in split_text_sentences(steps_text):
        if any(keyword in sentence for keyword in ("检查", "确认", "验证", "查看", "保持", "生效")):
            inferred.append(sentence)
    return unique_list(inferred)


def make_assertions(module_id: str, checkpoint_summary: list[str]) -> list[dict[str, Any]]:
    assertions: list[dict[str, Any]] = []
    marker = PAGE_REGISTRY.get(module_to_page_id(module_id), "")
    for index, item in enumerate(checkpoint_summary[:3]):
        if marker:
            assertions.append(
                {
                    "name": f"checkpoint_{index + 1}",
                    "type": "assert_text_presence",
                    "value": marker,
                    "params": {
                        "bundle_name": ADAPTER_CONFIG.bundle_name,
                        "present": True,
                        "timeout_ms": 1500,
                    },
                }
            )
            break
        assertions.append(
            {
                "name": f"checkpoint_{index + 1}",
                "type": "assert_text_presence",
                "value": item,
                "params": {
                    "bundle_name": ADAPTER_CONFIG.bundle_name,
                    "present": True,
                    "timeout_ms": 1000,
                },
            }
        )
    return assertions


def has_any_keyword(text: str, keywords: tuple[str, ...]) -> bool:
    return any(keyword in text for keyword in keywords)


def append_flow(flow_items: list[dict[str, Any]], flow_ref: str, params: dict[str, Any] | None = None) -> None:
    flow_items.append({"ref": flow_ref, "params": params or {}})


def infer_structured_flow(module_id: str, steps_text: str, checkpoints_text: str) -> tuple[list[dict[str, Any]], list[str]]:
    flow_items: list[dict[str, Any]] = []
    gap_items: list[str] = []

    if module_id not in {"bootstrap", "navigation"}:
        append_flow(flow_items, "app.launch", {"timeout_sec": 20})
    if module_id in MODULE_ID_TO_PAGE_ID:
        append_flow(flow_items, "navigation.open_page", {"page_id": module_to_page_id(module_id)})

    combined = f"{steps_text} {checkpoints_text}"

    if module_id == "tool-settings":
        if has_any_keyword(combined, ("启动时身份校验", "启动认证", "启动校验")):
            append_flow(flow_items, "entity.toggle", {"domain": "tool_settings", "entity": "startup_auth", "data": {"target_state": "invert"}})
        elif has_any_keyword(combined, ("认证方式", "PIN", "密码方式")):
            method = "pin" if "PIN" in combined.upper() else "password"
            append_flow(flow_items, "entity.update", {"domain": "tool_settings", "entity": "auth_method", "data": {"method": method}})
        elif has_any_keyword(combined, ("密码", "口令")):
            append_flow(flow_items, "tool_settings.set_password", {"new_password": "1234", "confirm_password": "1234"})
        else:
            gap_items.append("工具设置步骤未能稳定映射到具体 flow")

    elif module_id == "firewall":
        if has_any_keyword(combined, ("开启", "关闭", "开关", "状态切换", "切换")):
            append_flow(flow_items, "firewall.toggle_status", {"target_state": "invert"})
        if has_any_keyword(combined, ("规则", "白名单", "黑名单", "新增")):
            append_flow(flow_items, "entity.create", {"domain": "firewall", "entity": "rule", "variant": "domain", "data": {"name": "imported-domain-rule", "domain": "example.com", "direction": "out", "policy": "deny", "protocol": "tcp", "port": "443"}})
        if has_any_keyword(combined, ("删除", "移除")):
            append_flow(flow_items, "entity.delete", {"domain": "firewall", "entity": "rule", "variant": "domain", "data": {"name": "imported-domain-rule", "domain": "example.com"}})
        if has_any_keyword(combined, ("浏览器", "网址", "URL")):
            append_flow(flow_items, "browser.open_url", {"url": "https://example.com"})
        if len(flow_items) <= 2:
            gap_items.append("防火墙步骤未能稳定映射到具体 flow")

    elif module_id == "peripheral":
        if has_any_keyword(combined, ("接口", "蓝牙", "wifi", "USB 接口", "切换")):
            append_flow(flow_items, "entity.toggle", {"domain": "peripheral", "entity": "interface", "variant": "usb", "data": {"target_state": "invert"}})
        if has_any_keyword(combined, ("USB 存储", "存储策略", "只读", "禁用")):
            append_flow(flow_items, "entity.update", {"domain": "peripheral", "entity": "usb_storage_policy", "data": {"policy": "read_only"}})
        if has_any_keyword(combined, ("USB 白名单", "usb 白名单")):
            append_flow(flow_items, "entity.create", {"domain": "peripheral", "entity": "whitelist", "variant": "usb", "data": {"device_id": "USB-DEVICE-001"}})
        if has_any_keyword(combined, ("蓝牙白名单", "bluetooth")):
            append_flow(flow_items, "entity.create", {"domain": "peripheral", "entity": "whitelist", "variant": "bluetooth", "data": {"device_id": "BT-DEVICE-001"}})
        if has_any_keyword(combined, ("USB 黑名单", "usb 黑名单")):
            append_flow(flow_items, "entity.create", {"domain": "peripheral", "entity": "blacklist", "variant": "usb", "data": {"device_id": "USB-DEVICE-002"}})
        if len(flow_items) <= 2:
            gap_items.append("外设步骤未能稳定映射到具体 flow")

    elif module_id == "identity":
        if has_any_keyword(combined, ("口令", "密码复杂度", "密码策略")):
            append_flow(flow_items, "entity.update", {"domain": "identity", "entity": "password_policy", "data": {"min_length": "8"}})
        if has_any_keyword(combined, ("域", "账号策略", "domain")):
            append_flow(flow_items, "entity.update", {"domain": "identity", "entity": "domain_policy", "data": {"password_max_age_days": "90", "expiration_notify_days": "7", "auth_validity_minutes": "30"}})
        if len(flow_items) <= 2:
            gap_items.append("身份鉴别步骤未能稳定映射到具体 flow")

    elif module_id == "log-manage":
        if has_any_keyword(combined, ("导出", "导出日志")):
            append_flow(flow_items, "logs.export", {})
        if has_any_keyword(combined, ("策略变更", "触发日志", "修改策略")):
            append_flow(flow_items, "logs.change_any_policy", {})
        if len(flow_items) <= 2:
            gap_items.append("日志步骤未能稳定映射到具体 flow")

    elif module_id in {"dashboard", "navigation"}:
        if has_any_keyword(combined, ("主题", "菜单")):
            append_flow(flow_items, "theme_menu.open", {})

    else:
        gap_items.append("模块未识别，无法自动生成具体 flow")

    return normalize_flow_items(flow_items), unique_list(gap_items)


def transform_record(record: dict[str, Any]) -> dict[str, Any]:
    transformed = copy.deepcopy(record)
    module_id = transformed.get("module_id", "")
    steps_text = transformed.get("steps_text", "")
    checkpoints_text = transformed.get("checkpoints_text", "")
    flow_items, transform_gaps = infer_structured_flow(module_id, steps_text, checkpoints_text)
    checkpoint_summary = summarize_checkpoints(checkpoints_text, steps_text)
    assertions = make_assertions(module_id, checkpoint_summary)
    bridge_status = compute_bridge_status(flow_items, load_bridge_action_map())
    gap_status, gap_items = classify_gap_status(flow_items, assertions, bridge_status)
    transformed["structured_flow"] = flow_items
    transformed["structured_assertions"] = assertions
    transformed["checkpoint_summary"] = checkpoint_summary
    transformed["bridge_coverage_status"] = bridge_status
    transformed["capability_gap_status"] = gap_status if not transform_gaps else ("mixed" if gap_status != "none" else "flow_gap")
    transformed["capability_gap_items"] = unique_list(transform_gaps + gap_items)
    transformed["capability_gap_breakdown"] = build_gap_breakdown(transformed["capability_gap_items"])
    transformed["review_status"] = transformed.get("review_status") or "unreviewed"
    transformed["notes"] = unique_list(list(transformed.get("notes", [])))
    return transformed


def import_excel_workbook(data: bytes, filename: str = "upload.xlsx") -> ImportResult:
    ensure_directories()
    catalog = load_catalog()
    previous_records = {record.get("case_id"): copy.deepcopy(record) for record in catalog.get("records", [])}
    existing_records = {record.get("case_id"): record for record in catalog.get("records", [])}
    rows = parse_xlsx_rows(data)
    report = {
        "filename": filename,
        "imported_at": utc_now(),
        "total_rows": len(rows),
        "imported": [],
        "conflicts": [],
        "warnings": [],
        "diff": {"added": [], "updated": [], "unchanged": [], "conflicts": []},
        "summary": {},
    }

    next_records = list(catalog.get("records", []))
    used_ids = {record.get("case_id") for record in next_records}
    last_module_name = ""
    for index, row in enumerate(rows):
        case_name = resolve_header_value(row, "case_name")
        module_name = resolve_header_value(row, "module_name")
        if module_name:
            last_module_name = module_name
        elif last_module_name:
            module_name = last_module_name
        module_id = normalize_module_id(module_name)
        steps_text = resolve_header_value(row, "steps_text")
        checkpoints_text = resolve_header_value(row, "checkpoints_text")
        notes_text = resolve_header_value(row, "notes")
        enabled_text = resolve_header_value(row, "enabled").lower()
        case_id = resolve_header_value(row, "case_id")
        if not case_id:
            case_id = derive_case_id(module_id or "AUTO", index, case_name)
            while case_id in used_ids:
                case_id = derive_case_id(module_id or "AUTO", index + 100, case_name)

        if case_id in existing_records:
            conflict = {"row_index": index + 2, "case_id": case_id, "reason": "case_id 已存在台账"}
            report["conflicts"].append(conflict)
            report["diff"]["conflicts"].append(conflict)
            continue

        record = default_record(case_id)
        record.update(
            {
                "case_id": case_id,
                "case_name": case_name or f"导入记录 {index + 1}",
                "module_id": module_id,
                "module_name": module_name or module_id_to_name(module_id),
                "status": "planned" if enabled_text in {"n", "no", "否", "0"} else "draft",
                "source": "excel_import",
                "preconditions_text": resolve_header_value(row, "preconditions_text"),
                "steps_text": steps_text,
                "checkpoints_text": checkpoints_text,
                "notes": unique_list([notes_text]) if notes_text else [],
                "review_status": "unreviewed",
                "review_notes": [],
            }
        )
        if not module_id:
            record["capability_gap_status"] = "flow_gap"
            record["capability_gap_items"] = ["模块名称无法映射到系统模块 ID"]
            report["warnings"].append({"row_index": index + 2, "case_id": case_id, "reason": "模块名称未映射"})
        record = transform_record(record)
        next_records.append(record)
        used_ids.add(case_id)
        existing_records[case_id] = record
        report["imported"].append({"row_index": index + 2, "case_id": case_id, "module_id": module_id, "status": record["status"]})
        previous = previous_records.get(case_id)
        if previous is None:
            report["diff"]["added"].append({"case_id": case_id, "status": record["status"], "module_id": module_id})
        elif previous == record:
            report["diff"]["unchanged"].append({"case_id": case_id})
        else:
            changed_fields = sorted(key for key in record.keys() if previous.get(key) != record.get(key))
            report["diff"]["updated"].append({"case_id": case_id, "changed_fields": changed_fields})

    new_catalog = save_catalog({"version": CATALOG_VERSION, "updated_at": utc_now(), "records": next_records})
    report["summary"] = {
        "added": len(report["diff"]["added"]),
        "updated": len(report["diff"]["updated"]),
        "unchanged": len(report["diff"]["unchanged"]),
        "conflicts": len(report["diff"]["conflicts"]),
        "warnings": len(report["warnings"]),
    }
    save_json(IMPORT_REPORT_PATH, report)
    return ImportResult(catalog=new_catalog, report=report)


def import_excel_path(path: Path) -> ImportResult:
    return import_excel_workbook(path.read_bytes(), filename=path.name)


def transform_catalog_records() -> dict[str, Any]:
    catalog = load_catalog()
    transformed_records = [transform_record(record) for record in catalog.get("records", [])]
    updated = save_catalog({"version": CATALOG_VERSION, "updated_at": utc_now(), "records": transformed_records})
    build_coverage_snapshot(updated)
    return updated


def build_case_payload_from_record(record: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "case_id": record.get("case_id", ""),
        "case_name": record.get("case_name", ""),
        "module": record.get("module_id", ""),
        "preconditions": [],
        "flow": normalize_flow_items(record.get("structured_flow", [])),
        "assertions": record.get("structured_assertions", []),
        "result_policy": {"allow_unknown": True, "stop_on_failure": True},
        "notes": unique_list(record.get("notes", []) + ["Generated from test asset catalog; pending manual review."]),
    }
    validate_case_contract(payload)
    return payload


def module_id_to_case_dir(module_id: str) -> Path:
    mapping = {
        "dashboard": CASES_DIR / "navigation",
        "firewall": CASES_DIR / "firewall",
        "identity": CASES_DIR / "identity",
        "log-manage": CASES_DIR / "logs",
        "navigation": CASES_DIR / "navigation",
        "peripheral": CASES_DIR / "peripheral",
        "tool-settings": CASES_DIR / "tool_settings",
        "bootstrap": CASES_DIR / "smoke",
    }
    return mapping.get(module_id, CASES_DIR / "navigation")


def promote_case_draft(case_id: str) -> dict[str, Any]:
    ensure_directories()
    catalog = load_catalog()
    next_records: list[dict[str, Any]] = []
    promoted: dict[str, Any] | None = None

    for record in catalog.get("records", []):
        updated_record = copy.deepcopy(record)
        if updated_record.get("case_id") != case_id:
            next_records.append(updated_record)
            continue

        if not updated_record.get("module_id"):
            raise ValueError("selected record is missing module_id")
        if updated_record.get("status") not in {"draft", "planned"}:
            raise ValueError("only draft or planned records can be promoted")
        if not updated_record.get("structured_flow"):
            raise ValueError("structured_flow is empty; promotion is blocked")

        payload = build_case_payload_from_record(updated_record)
        target_dir = module_id_to_case_dir(updated_record["module_id"])
        target_dir.mkdir(parents=True, exist_ok=True)
        target_path = target_dir / f"{updated_record['case_id'].lower()}.json"
        save_json(target_path, payload)

        updated_record["case_path"] = to_posix(target_path)
        updated_record["status"] = "implemented"
        promoted = {
            "case_id": updated_record["case_id"],
            "case_path": updated_record["case_path"],
            "status": updated_record["status"],
        }
        next_records.append(updated_record)

    if promoted is None:
        raise ValueError(f"case_id not found: {case_id}")

    updated_catalog = save_catalog({"version": CATALOG_VERSION, "updated_at": utc_now(), "records": next_records})
    build_coverage_snapshot(updated_catalog)
    return {"promoted_at": utc_now(), "record": promoted, "catalog": updated_catalog}


def batch_promote_case_drafts(case_ids: list[str]) -> dict[str, Any]:
    requested = unique_list(case_ids)
    promoted: list[dict[str, Any]] = []
    blocked: list[dict[str, str]] = []

    for case_id in requested:
        try:
            result = promote_case_draft(case_id)
            promoted.append(result["record"])
        except Exception as exc:  # noqa: BLE001
            blocked.append({"case_id": case_id, "reason": str(exc)})

    return {
        "processed_at": utc_now(),
        "requested_count": len(requested),
        "promoted_count": len(promoted),
        "blocked_count": len(blocked),
        "promoted": promoted,
        "blocked": blocked,
        "catalog": load_catalog(),
    }


def batch_delete_catalog_records(case_ids: list[str]) -> dict[str, Any]:
    requested = unique_list(case_ids)
    if not requested:
        return {
            "processed_at": utc_now(),
            "requested_count": 0,
            "deleted_count": 0,
            "deleted": [],
            "blocked": [],
            "catalog": load_catalog(),
        }

    catalog = load_catalog()
    remaining_records: list[dict[str, Any]] = []
    deleted: list[dict[str, str]] = []
    blocked: list[dict[str, str]] = []

    for record in catalog.get("records", []):
        case_id = record.get("case_id", "")
        if case_id not in requested:
            remaining_records.append(record)
            continue

        if record.get("case_path"):
            blocked.append({"case_id": case_id, "reason": "正式 case 不能从页面直接删除"})
            remaining_records.append(record)
            continue

        deleted.append({"case_id": case_id, "status": record.get("status", "")})

    updated_catalog = save_catalog({"version": CATALOG_VERSION, "updated_at": utc_now(), "records": remaining_records})
    build_coverage_snapshot(updated_catalog)
    return {
        "processed_at": utc_now(),
        "requested_count": len(requested),
        "deleted_count": len(deleted),
        "blocked_count": len(blocked),
        "deleted": deleted,
        "blocked": blocked,
        "catalog": updated_catalog,
    }


def validate_test_assets() -> dict[str, Any]:
    catalog = load_catalog()
    errors: list[dict[str, str]] = []
    warnings: list[dict[str, str]] = []
    known_case_ids: set[str] = set()

    for record in catalog.get("records", []):
        case_id = record.get("case_id", "")
        if not case_id:
            errors.append({"case_id": "", "reason": "记录缺少 case_id"})
            continue
        if case_id in known_case_ids:
            errors.append({"case_id": case_id, "reason": "台账中存在重复 case_id"})
        known_case_ids.add(case_id)

        if record.get("module_id") and record.get("module_id") not in MODULE_ID_TO_NAME:
            warnings.append({"case_id": case_id, "reason": "module_id 未在已知模块映射内"})
        for suite_name in record.get("suite_membership", []):
            if suite_name not in SUITES:
                errors.append({"case_id": case_id, "reason": f"非法 suite 归属: {suite_name}"})
        for flow_item in record.get("structured_flow", []):
            flow_ref = flow_item.get("ref", "")
            if flow_ref not in FLOW_REGISTRY:
                errors.append({"case_id": case_id, "reason": f"未知 flow_ref: {flow_ref}"})
            for issue in template_support_issues(flow_ref, flow_item.get("params", {})):
                errors.append({"case_id": case_id, "reason": issue})

    report = {
        "validated_at": utc_now(),
        "error_count": len(errors),
        "warning_count": len(warnings),
        "errors": errors,
        "warnings": warnings,
    }
    save_json(VALIDATION_REPORT_PATH, report)
    return report


def upsert_record(record: dict[str, Any]) -> dict[str, Any]:
    catalog = load_catalog()
    records = catalog.get("records", [])
    case_id = record.get("case_id", "")
    if not case_id:
        raise ValueError("case_id is required")
    transformed = transform_record(record)
    replaced = False
    next_records: list[dict[str, Any]] = []
    for current in records:
        if current.get("case_id") == case_id:
            next_records.append(transformed)
            replaced = True
        else:
            next_records.append(current)
    if not replaced:
        next_records.append(transformed)
    return save_catalog({"version": CATALOG_VERSION, "updated_at": utc_now(), "records": next_records})


def parse_args_with_catalog(description: str) -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=description)
    parser.add_argument("--rebuild", action="store_true", help="Rebuild catalog before running this command when applicable.")
    return parser


def decode_base64_file(content_base64: str) -> bytes:
    return base64.b64decode(content_base64.encode("utf-8"))
